import openpyxl as px
from openpyxl.styles import PatternFill
import glob
import tkinter as tk

fill1 = PatternFill(patternType="solid", fgColor="C5D9F1")
fill2 = PatternFill(patternType="solid", fgColor="FDE9D9")

def hiyari():
    wb1 = px.Workbook()
    ws1 = wb1['Sheet']
    ws1.cell(row=1, column=1).value = "分類1"
    ws1.cell(row=1, column=2).value = "分類2"
    ws1.cell(row=1, column=3).value = "所属"
    ws1.cell(row=1, column=4).value = "社員番号"
    ws1.cell(row=1, column=5).value = "氏名"
    ws1.cell(row=1, column=6).value = "いつ年"
    ws1.cell(row=1, column=7).value = "いつ月"
    ws1.cell(row=1, column=8).value = "いつ日"
    ws1.cell(row=1, column=9).value = "要約"
    ws1.cell(row=1, column=10).value = "なぜ起きた"
    ws1.cell(row=1, column=11).value = "対策"
    ws1.cell(row=1, column=12).value = "対応状況"
    ws1.cell(row=1, column=13).value = "水平展開"
    ws1.cell(row=1, column=14).value = "コメント"
    ws1.cell(row=1, column=15).value = "ファイル名"
    ws1.cell(row=1, column=16).value = "エリア会共有"
    ws1.cell(row=1, column=17).value = "コメント（部長）"
    ws1.cell(row=1, column=18).value = "環安会共有"

    files = glob.glob('hiyari01/*.xlsx')
    i = 2

    for file in files:
        wb2 = px.load_workbook(filename=file, data_only=True)
        for ws2 in wb2:
            ws1.cell(row=i, column=1).value = ws2.cell(row=10, column=1).value  # 分類１
            ws1.cell(row=i, column=2).value = ws2.cell(row=10, column=3).value  # 分類2
            ws1.cell(row=i, column=3).value = ws2.cell(row=10, column=5).value  # 所属
            ws1.cell(row=i, column=4).value = ws2.cell(row=10, column=8).value  # 社員番号
            ws1.cell(row=i, column=5).value = ws2.cell(row=10, column=10).value  # 氏名
            ws1.cell(row=i, column=6).value = ws2.cell(row=12, column=3).value  # いつ年
            ws1.cell(row=i, column=7).value = ws2.cell(row=12, column=5).value  # いつ月
            ws1.cell(row=i, column=8).value = ws2.cell(row=12, column=7).value  # いつ日
            ws1.cell(row=i, column=9).value = ws2.cell(row=15, column=3).value  # 要約
            ws1.cell(row=i, column=10).value = ws2.cell(row=18, column=3).value  # なぜ起きた
            ws1.cell(row=i, column=11).value = ws2.cell(row=19, column=3).value  # 対策
            ws1.cell(row=i, column=12).value = ws2.cell(row=22, column=3).value  # 対応状況
            ws1.cell(row=i, column=13).value = ws2.cell(row=22, column=7).value  # 水平展開
            ws1.cell(row=i, column=14).value = ws2.cell(row=23, column=3).value  # コメント
            ws1.cell(row=i, column=15).value = file  # ファイル名
            i = i + 1

    sheet_max = ws1.max_row + 1  # シートの最終行を取得
    for j in reversed(range(sheet_max)):  # 最終行から逆ループ
        if j == 0:
            break
        if ws1.cell(row=j, column=1).value is None:  # A列が空白だったら
            ws1.delete_rows(j)  # 行削除

    for m in range(1, 15):
        ws1.cell(row=1, column=m).fill = fill1

    for m in range(16, 18):
        ws1.cell(row=1, column=m).fill = fill2

    wb1.save('hiyari02/summary.xlsx')
    statusbar["text"] = "Done"


root = tk.Tk()  # ウインドウの作成
root.geometry("350x100")  # ウインドウのサイズ指定

run_button = tk.Button(root, text="Run", command=hiyari)  # Runボタン設置
run_button.place(x=160, y=40)

statusbar = tk.Label(root, text="Ready", bd=1, relief=tk.SUNKEN, anchor=tk.W)  # ステータスバー設置
statusbar.pack(side=tk.BOTTOM, fill=tk.X)

root.mainloop()  # ウインドウ状態の維持
